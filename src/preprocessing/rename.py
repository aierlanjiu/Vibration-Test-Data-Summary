#如果有多文件组合的需求可以参考此程序
import openpyxl
import re
import os
from openpyxl import Workbook, load_workbook    # 导入所需库
import pandas as pd



# 文件夹路径
folder_path = 'root\\data\\raw_data'

# 文件名列表
file_list = ['9.62 order run down.xlsx', '9.62 order run up.xlsx', '26 order run up.xlsx', '26 order run down.xlsx']
file_list_frames = ['9.62 order run down_frames.xlsx', '9.62 order run up_frames.xlsx', '26 order run up_frames.xlsx', '26 order run down_frames.xlsx']
# 记录数字部分最大的数字
max_num_dict = {}

# 遍历文件名列表
for file_name in file_list:

    # 获取文件路径
    file_path = os.path.join(folder_path, file_name)
    # 打开工作簿
    wb = openpyxl.load_workbook(file_path)
    # 遍历所有工作表
    for sheet in wb.sheetnames:
        if sheet == "Sheet":  # 如果该工作表名称为"sheet"
            wb.remove(wb['Sheet'])
        else:         # 跳过该工作表，遍历下一个工作表
            num = re.findall('/d+', sheet)[0]
            # 更新数字部分最大的数字
            if file_name[:-5] not in max_num_dict or int(num) > max_num_dict[file_name[:-5]]:
                max_num_dict[file_name[:-5]] = int(num)
    # 关闭工作簿
    wb.close()

# 首先创建一个新的字典，用于存储修改后的键值对
new_dict = {}
for old_key in max_num_dict:
    # 获取旧键所对应的值
    value = max_num_dict[old_key]
    # 修改旧键名并将旧键值赋给新键
    new_key = old_key + '_frames'
    new_dict[new_key] = value

# 使用新的字典替换旧的字典
max_num_dict = new_dict
print(max_num_dict)

# 依次重命名工作表
for i, file_name in enumerate(file_list_frames):
    # 获取文件路径
    file_path = os.path.join(folder_path, file_name)
    # 打开工作簿
    wb = openpyxl.load_workbook(file_path)
    # 遍历所有工作表
    for sheet in wb.sheetnames:
        if sheet == "Sheet":  # 如果该工作表名称为"sheet"
            wb.remove(wb[sheet])
        else:
            # 获取工作表对象
            ws = wb[sheet]
            # 获取工作表名称中的数字部分
            num = re.findall('/d+', sheet)[0]
            # 将数字部分增加数字部分最大的数字
            new_num = int(num) + max_num_dict[file_name[:-5]]
            print(new_num)
            # 将工作表重命名为新名称
            ws.title = re.sub(num, str(new_num), sheet)
            print(ws.title)

    wb.save(file_name)  # 保存修改后的Excel文件
    # 关闭工作簿
    wb.close()
    # 保存工作簿
    wb.save(file_path)



# 遍历每个文件路径并逐一进行Excel写入操作
for i in range(len(file_list)):
    print(f"正在写入 {file_list[i]} 文件...")
    # 使用 read_excel() 函数读取原始Excel文件，获取数据框字典
    df_dict = pd.read_excel(file_list[i], sheet_name=None)
    df_dict_frames = pd.read_excel(file_list_frames[i],sheet_name=None)

    with pd.ExcelWriter(file_list[i]) as writer:
        # 遍历数据框字典，并依次将每个数据框写入新的工作表
        for sheet_name, df in df_dict.items():
            # 使用 to_excel() 方法将当前数据框写入新的文件中
            df.to_excel(writer, sheet_name=sheet_name)
        # 读取当前文件对应的 frames 文件，并将其中的数据写入到末尾
        print(f"正在读取 {file_list_frames[i]} 文件...")
        for sheet_name, df in df_dict_frames.items():
            print(f"正在写入 {sheet_name} 到Excel file...")
            # 使用 to_excel() 方法将当前数据框写入新的文件中
            df.to_excel(writer, sheet_name=sheet_name)



    print(f"{file_list[i]} 文件已经更新完毕！")


